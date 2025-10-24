"""
Excel export utilities using openpyxl.

These functions allow exporting either a single patient's data (across all
tables) or the entire database into an Excel workbook with multiple sheets.
The workbook will contain a sheet per table with column headers.  For a
single patient export, only the rows associated with that patient are included.

Requires openpyxl; ensure it is installed prior to packaging.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook

from db.models import Database
# 引入日期格式化函数
from utils.validators import format_date6, format_birth_ym4, format_birth_ym6
# 引入分期查询函数
# 已删除临床分期映射功能，不再导入 staging.lookup 中的分期函数

# Fields that should not appear in exported files.  These columns either come from
# deprecated UI elements (e.g. vendor_lab) or have been removed from the
# current version of the application (e.g. lymph node totals).
EXCLUDE_FIELDS = {"vendor_lab", "ln_total", "ln_positive", "patient_id"}

# 与 CSV 导出一致的日期字段集合
DATE_FIELDS_8 = {
    "surgery_date6",
    "report_date",
    "test_date",
    "last_visit_date",
    "death_date",
    "relapse_date",
}

def _format_value(col: str, val: object) -> object:
    """根据列名格式化日期字段。"""
    if val is None or val == "":
        return val
    s = str(val)
    if col == "birth_ym4":
        if len(s) == 4 and s.isdigit():
            formatted = format_birth_ym4(s)
            return formatted.replace("-", "") if formatted else s
        if len(s) == 6 and s.isdigit():
            formatted = format_birth_ym6(s)
            return formatted.replace("-", "") if formatted else s
        return s
    if col in DATE_FIELDS_8:
        if len(s) == 6 and s.isdigit():
            formatted = format_date6(s)
            return formatted.replace("-", "") if formatted else s
        return s
    return val

def _format_row_dates(row_dict: dict) -> dict:
    new_row = {}
    for k, v in row_dict.items():
        if k in EXCLUDE_FIELDS:
            continue
        new_row[k] = _format_value(k, v)
    return new_row

# 移除 `_annotate_stage` 函数。临床分期映射功能已取消，不再补充分期字段。


def _clean_row(row_dict: dict) -> dict:
    """Remove unwanted fields and format date values before export."""
    formatted = _format_row_dates(row_dict)
    return {k: v for k, v in formatted.items() if k not in EXCLUDE_FIELDS}


def _reorder_pathology(row_dict: dict) -> dict:
    """Reorder pathology row so that airway_spread appears before pleural_invasion.

    When exporting Pathology data, the airway_spread column should precede
    pleural_invasion.  This helper takes a cleaned row dictionary and
    inserts the airway_spread key before pleural_invasion if both are present.
    """
    # Only reorder when both keys exist and airway_spread comes later in order
    if "airway_spread" in row_dict and "pleural_invasion" in row_dict:
        # Extract the value and delete the key to reposition
        airway_val = row_dict.pop("airway_spread")
        new_dict = {}
        for k, v in row_dict.items():
            # Insert airway_spread just before pleural_invasion
            if k == "pleural_invasion":
                new_dict["airway_spread"] = airway_val
            new_dict[k] = v
        return new_dict
    return row_dict


def _write_sheet(wb: Workbook, sheet_name: str, rows: Iterable[dict]) -> None:
    """写入工作表数据，包含异常处理。"""
    try:
        ws = wb.create_sheet(title=sheet_name)
        # Convert to list, remove excluded fields and format dates
        rows_list = []
        for row in rows:
            try:
                row_dict = dict(row) if not isinstance(row, dict) else row
                cleaned = _clean_row(row_dict)
                # For Pathology sheet reorder airway_spread before pleural_invasion
                if sheet_name == "Pathology":
                    cleaned = _reorder_pathology(cleaned)
                rows_list.append(cleaned)
            except Exception as e:
                print(f"Warning: Failed to process row in {sheet_name}: {e}")
                continue
        
        if not rows_list:
            return
        
        # Write header based on cleaned row keys (order matters)
        header = list(rows_list[0].keys())
        ws.append(header)
        for row in rows_list:
            # Convert sqlite3.Row to dict to support .get() method
            row_dict = dict(row)
            ws.append([row_dict.get(col) for col in header])
    except Exception as e:
        print(f"Error: Failed to write sheet {sheet_name}: {e}")
        raise


def export_patient_to_excel(db: Database, patient_id: int, file_path: Path) -> None:
    """Export data for a single patient to an Excel file."""
    try:
        wb = Workbook()
        # Remove the default sheet created by openpyxl
        wb.remove(wb.active)
        # Fetch rows per table
        tables = ["Patient", "Surgery", "Pathology", "Molecular", "FollowUpEvent"]
        # Fetch the patient once to obtain hospital_id.  Convert sqlite3.Row to dict
        patient_row = db.get_patient_by_id(patient_id)
        
        if not patient_row:
            raise ValueError(f"Patient with ID {patient_id} not found")
        
        # patient_dict_list is a list of dicts to feed into sheet writer
        patient_dict_list: List[dict] = []
        # Determine hospital_id from the converted dict
        hospital_id = None
        
        pr_dict = dict(patient_row)
        # 不再补充分期字段，直接使用患者行内容
        patient_dict_list = [pr_dict]
        hospital_id = pr_dict.get("hospital_id")
        
        for table in tables:
            try:
                if table == "Patient":
                    rows = patient_dict_list
                else:
                    # Many-to-one tables
                    if table == "Surgery":
                        items = db.get_surgeries_by_patient(patient_id)
                    elif table == "Pathology":
                        items = db.get_pathologies_by_patient(patient_id)
                    elif table == "Molecular":
                        items = db.get_molecular_by_patient(patient_id)
                    else:
                        items = []
                    rows = []
                    for row in items:
                        rdict = dict(row)
                        if hospital_id is not None:
                            rdict = {"hospital_id": hospital_id, **rdict}
                        rows.append(rdict)
                _write_sheet(wb, table, rows)
            except Exception as e:
                print(f"Warning: Failed to export table {table}: {e}")
                continue
        
        # 确保目标目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        
    except PermissionError as e:
        raise PermissionError(f"无法写入文件 {file_path}，请检查文件权限或文件是否被占用") from e
    except Exception as e:
        raise Exception(f"导出Excel文件失败: {str(e)}") from e


def export_all_to_excel(db: Database, file_path: Path) -> None:
    """Export entire database to an Excel file (one sheet per table)."""
    try:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        # Precompute mapping from patient_id to hospital_id
        patient_rows = db.export_table("Patient")
        pat_map = {}
        for row in patient_rows:
            rdict = dict(row)
            pat_map[rdict.get("patient_id")] = rdict.get("hospital_id")
        
        # For each table, fetch all rows and attach hospital_id for non-Patient tables
        for table in ["Patient", "Surgery", "Pathology", "Molecular", "FollowUpEvent"]:
            try:
                rows = db.export_table(table)
                rows_dicts: List[dict] = []
                for row in rows:
                    rdict = dict(row)
                    if table == "Patient":
                        # 不再补充分期字段。患者行直接使用原字段
                        pass
                    else:
                        pid = rdict.get("patient_id")
                        if pid in pat_map:
                            rdict = {"hospital_id": pat_map[pid], **rdict}
                    rows_dicts.append(rdict)
                _write_sheet(wb, table, rows_dicts)
            except Exception as e:
                print(f"Warning: Failed to export table {table}: {e}")
                continue
        
        # 确保目标目录存在
        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(file_path)
        
    except PermissionError as e:
        raise PermissionError(f"无法写入文件 {file_path}，请检查文件权限或文件是否被占用") from e
    except Exception as e:
        raise Exception(f"导出Excel文件失败: {str(e)}") from e

